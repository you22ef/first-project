from tkinter import*
from tkinter import ttk
import openpyxl as xl
root=Tk()#creat a window 
root.geometry('1000x600+280+100')#the size of the window 
style=ttk.Style()
style.theme_use("classic") #the style of the window 
L1=Label(root,text="Population app",font=('Helvatical bold',50,'underline'),fg='red').place(x=275,y=0)
L2=Label(root,text="Please enter the name of country",fg='blue',font=('Helvatical bold',40)).place(x=129,y=130)
def country():
    global L3
    global L4
    global L5
    try:
        user=b.get() #get the name of country
        print("\n")
        wb = xl.load_workbook(f"{user}.xlsx") #get the name of country
        
        sheet = wb["Sheet1"] #get the name of sheet 

        total = 0
        
        maxx=[]
        cities = []
        
        for r in range (1,sheet.max_row+1 ) :
                
                cell = sheet.cell(r,1)
                
                print(f"{cell.value}:",end="")
                
                cities.append(cell.value)#get the names of countries
                
                cell = sheet.cell(r,2)  
                
                print(f"{cell.value}")
                total+=cell.value
                maxx.append(cell.value)#get the numbers of population
        L6=  Label(root,text="invalid input",fg='#f0f0f0',font=('Helvatical bold',20)).place(x=400,y=205,height=40,width=200)      
        L3=  Label(root,text=f"\ntotal poulation :{total}\n",font=('Helvatical bold',30),fg='blue').place(x=20,y=290)          
        L4=  Label(root,text=f"the highest population is in {cities[maxx.index(max(maxx))]} : {max(maxx)}\n",font=('Helvatical bold',30),fg='green').place(x=20,y=420) 
        L5=  Label(root,text=f"the minimum population is in {cities[maxx.index(min(maxx))]} : {min(maxx)}\n",font=('Helvatical bold',30),fg='red').place(x=20,y=500)       
    except:
        L3=  Label(root,text=f"total poulation ",font=('Helvatical bold',30),fg='#f0f0f0').place(x=20,y=290,width=1000,height=1000)  
        L4=  Label(root,text=f"the highest population ",font=('Helvatical bold',30),fg='#f0f0f0').place(x=20,y=420,width=1000) 
        L5=  Label(root,text=f"the minimum population is in",font=('Helvatical bold',30),fg='#f0f0f0').place(x=20,y=500,width=1000,height=1000)       
        L6= Label(root,text="invalid input",bg='yellow',fg='red',font=('Helvatical bold',20)).place(x=400,y=205,height=40,width=200)
                
Button1=Button(text='Enter',fg='red',bg='yellow',font=('Helvatical bold',25),command=country,cursor='hand2').place(x=710,y=245,width=130,height=45)#creat a button 
b=StringVar()        
text=Entry(textvariable=b).place(x=310,y=250,height=40,width=370)#entry to enter the name of country




root.mainloop()