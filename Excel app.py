import openpyxl as xl 
try:
    user=str(input("\n\nplease enter the name of country: "))#get the name of country
    print("\n")
    wb = xl.load_workbook(f"{user}.xlsx")#get the name of country
    
    sheet = wb["Sheet1"] #get the name of sheet

    total = 0
    
    maxx=[]
    cities = []
    
    for r in range (1,sheet.max_row+1 ) :
            
              cell = sheet.cell(r,1)
              
              print(f"{cell.value}:",end="")
              
              cities.append(cell.value)#get the names of countries
              
              cell = sheet.cell(r,2)  
              
              print(f"{cell.value}")
              total+=cell.value
              maxx.append(cell.value)#get the numbers of population
    
    print(f"\ntotal :{total}\n")  
    print(f"the highest population is in {cities[maxx.index(max(maxx))]} : {max(maxx)}\n")
    print(f"the minimum population is in {cities[maxx.index(min(maxx))]} : {min(maxx)}\n")        
except:
    print("invalid input")    